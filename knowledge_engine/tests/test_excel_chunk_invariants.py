# SPDX-FileCopyrightText: 2026 Weibo, Inc.
#
# SPDX-License-Identifier: Apache-2.0

"""Invariant tests for the Excel ingestion path.

These do not mirror individual review findings; they pin the structural
properties that every future change must preserve:

1. Row atomicity: a chunk never cuts a serialized row or cell fragment,
   for any workbook content and any budget at or above the minimum.
2. Dual-form budget: retrieval text and display text (prefixes included)
   each stay within the configured chunk size.
3. Reader/splitter agreement: whatever extension the indexer was
   constructed with, a real Excel file is chunked by the row-atomic
   splitter, never by a generic splitter that would shred the canonical
   row JSON.
"""

from __future__ import annotations

import io
import json
import random
from unittest.mock import MagicMock

from llama_index.core import Document
from openpyxl import Workbook

from knowledge_engine.excel import parse_excel_fragment_line, parse_excel_row_line
from knowledge_engine.index.indexer import DocumentIndexer
from knowledge_engine.splitter.excel_rows import (
    MIN_EXCEL_CHUNK_BUDGET,
    build_excel_hierarchical_nodes,
    build_excel_row_nodes,
)
from knowledge_engine.storage.chunk_metadata import ChunkMetadata

_VALUE_ALPHABET = 'ab中文🎉;:\n"\\'


def _parse_canonical_line(line: str) -> None:
    """One chunk display line must be a whole row or a whole fragment."""
    if '"fragment_index":' in line:
        parse_excel_fragment_line(line)
    else:
        parse_excel_row_line(line)


def _random_canonical_document(rng: random.Random) -> Document:
    """Build a canonical row document exercising adversarial cell values."""
    lines = []
    for source_row in range(1, rng.randint(2, 15) + 1):
        cells = [
            [column, value]
            for column in sorted(rng.sample(range(1, 30), rng.randint(1, 6)))
            for value in [
                rng.choice(
                    [
                        rng.randint(-(10**6), 10**6),
                        rng.choice([True, False]),
                        "".join(
                            rng.choice(_VALUE_ALPHABET)
                            for _ in range(rng.randint(0, 500))
                        ),
                        {"formula": "=SUM(A:A)", "value": rng.randint(0, 99)},
                    ]
                )
            ]
        ]
        lines.append(
            json.dumps(
                {"source_row": source_row, "cells": cells},
                ensure_ascii=False,
                separators=(",", ":"),
                allow_nan=False,
            )
        )
    sheet_name = rng.choice(["Data", "", "带换行\n的名字", "很" * 60, "Sheet 3"])
    return Document(text="\n".join(lines), metadata={"sheet_name": sheet_name})


def _assert_row_atomic_and_bounded(nodes, budget: int) -> None:
    for node in nodes:
        retrieval = node.metadata["retrieval_text"]
        display = node.metadata["display_text"]
        # Prefixes are paid out of the budget, not appended on top.
        assert len(retrieval) <= budget, retrieval
        assert len(display) <= budget, display
        # Every display line is a complete canonical row or fragment.
        for line in node.text.split("\n"):
            _parse_canonical_line(line)


def test_any_workbook_any_budget_keeps_rows_atomic_and_bounded() -> None:
    """Row atomicity and dual-form budgets hold across random workbooks."""
    rng = random.Random(20260818)
    for _ in range(50):
        document = _random_canonical_document(rng)
        budget = rng.choice([MIN_EXCEL_CHUNK_BUDGET, 129, 160, 256, 512, 1024, 4096])

        nodes = build_excel_row_nodes(documents=[document], chunk_size=budget)
        _assert_row_atomic_and_bounded(nodes, budget)

        parent_budget = rng.choice([budget, 2 * budget])
        result = build_excel_hierarchical_nodes(
            documents=[document],
            parent_chunk_size=parent_budget,
            child_chunk_size=budget,
        )
        _assert_row_atomic_and_bounded(result.parent_nodes, parent_budget)
        _assert_row_atomic_and_bounded(result.child_nodes, budget)


def _xlsx_bytes(*, long_cell: bool) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "数据"
    for row_index in range(8):
        value = "长" * 3000 if long_cell and row_index == 3 else f"r{row_index}"
        sheet.cell(row=row_index + 1, column=1, value=value)
        sheet.cell(row=row_index + 1, column=2, value=f"c{row_index}")
    payload = io.BytesIO()
    workbook.save(payload)
    workbook.close()
    return payload.getvalue()


def _index_xlsx(*, long_cell: bool, constructor_extension: str | None) -> list:
    storage_backend = MagicMock()
    storage_backend.index_with_metadata.return_value = {"indexed_count": 1}
    indexer = DocumentIndexer(
        storage_backend=storage_backend,
        embed_model=MagicMock(),
        splitter_config={"chunk_strategy": "flat", "format_enhancement": "none"},
        file_extension=constructor_extension,
    )
    indexer.index_from_binary(
        binary_data=_xlsx_bytes(long_cell=long_cell),
        file_extension=".xlsx",
        chunk_metadata=ChunkMetadata(
            knowledge_id="1",
            doc_ref="data",
            source_file="data.xlsx",
            created_at="2026-08-18T00:00:00+00:00",
        ),
    )
    return storage_backend.index_with_metadata.call_args.kwargs["nodes"]


def test_real_excel_extension_wins_over_constructor_extension() -> None:
    """Reader and splitter must agree: a real .xlsx is row-atomic even
    when the indexer was constructed for another extension."""
    nodes = _index_xlsx(long_cell=True, constructor_extension=".txt")

    assert nodes, "indexing must produce nodes"
    for node in nodes:
        # The generic splitter would shred the canonical row JSON into
        # fragments that no longer parse; row-atomic output always does.
        for line in node.text.split("\n"):
            _parse_canonical_line(line)
        assert "retrieval_text" in node.metadata
        assert "display_text" in node.metadata


def test_real_excel_path_extension_wins_over_constructor_extension(
    tmp_path,
) -> None:
    """Same agreement for file-path indexing with a reused indexer."""
    path = tmp_path / "book.xlsx"
    path.write_bytes(_xlsx_bytes(long_cell=False))

    storage_backend = MagicMock()
    storage_backend.index_with_metadata.return_value = {"indexed_count": 1}
    indexer = DocumentIndexer(
        storage_backend=storage_backend,
        embed_model=MagicMock(),
        splitter_config={"chunk_strategy": "flat", "format_enhancement": "none"},
        file_extension=".txt",
    )
    indexer.index_document(
        file_path=str(path),
        chunk_metadata=ChunkMetadata(
            knowledge_id="1",
            doc_ref="book",
            source_file="book.xlsx",
            created_at="2026-08-18T00:00:00+00:00",
        ),
    )

    nodes = storage_backend.index_with_metadata.call_args.kwargs["nodes"]
    assert nodes, "indexing must produce nodes"
    for node in nodes:
        for line in node.text.split("\n"):
            _parse_canonical_line(line)
        assert "retrieval_text" in node.metadata
        assert "display_text" in node.metadata


def test_uppercase_constructor_extension_still_dispatches_excel(
    tmp_path,
) -> None:
    """The resolved extension normalizes case, so a suffixless file declared
    Excel by an uppercase constructor default is still read row-atomically;
    before the fix the reader saw non-Excel while the splitter saw Excel."""
    path = tmp_path / "book"  # no suffix; the constructor default decides
    path.write_bytes(_xlsx_bytes(long_cell=False))

    storage_backend = MagicMock()
    storage_backend.index_with_metadata.return_value = {"indexed_count": 1}
    indexer = DocumentIndexer(
        storage_backend=storage_backend,
        embed_model=MagicMock(),
        splitter_config={"chunk_strategy": "flat", "format_enhancement": "none"},
        file_extension=".XLSX",
    )
    indexer.index_document(
        file_path=str(path),
        chunk_metadata=ChunkMetadata(
            knowledge_id="1",
            doc_ref="book",
            source_file="book",
            created_at="2026-08-18T00:00:00+00:00",
        ),
    )

    nodes = storage_backend.index_with_metadata.call_args.kwargs["nodes"]
    assert nodes, "indexing must produce nodes"
    for node in nodes:
        for line in node.text.split("\n"):
            _parse_canonical_line(line)
        assert node.metadata["sheet_name"] == "数据"


def test_suffixless_excel_file_stays_row_atomic_with_constructor_default(
    tmp_path,
) -> None:
    """A suffixless file declared Excel by the indexer constructor must be
    read by the structure-aware reader and chunked row-atomically; before
    the single-point dispatch fix the reader fell back to raw bytes and
    the strict splitter crashed on binary garbage."""
    path = tmp_path / "book"  # no suffix
    path.write_bytes(_xlsx_bytes(long_cell=True))

    storage_backend = MagicMock()
    storage_backend.index_with_metadata.return_value = {"indexed_count": 1}
    indexer = DocumentIndexer(
        storage_backend=storage_backend,
        embed_model=MagicMock(),
        splitter_config={"chunk_strategy": "flat", "format_enhancement": "none"},
        file_extension=".xlsx",
    )
    indexer.index_document(
        file_path=str(path),
        chunk_metadata=ChunkMetadata(
            knowledge_id="1",
            doc_ref="book",
            source_file="book",
            created_at="2026-08-18T00:00:00+00:00",
        ),
    )

    nodes = storage_backend.index_with_metadata.call_args.kwargs["nodes"]
    assert nodes, "indexing must produce nodes"
    for node in nodes:
        for line in node.text.split("\n"):
            _parse_canonical_line(line)
        # File-level metadata survives sanitize_metadata for the reader-
        # direct branch, keeping one metadata shape across reader kinds.
        assert node.metadata["sheet_name"]
        assert node.metadata["file_name"]
        assert int(node.metadata["file_size"]) > 0
