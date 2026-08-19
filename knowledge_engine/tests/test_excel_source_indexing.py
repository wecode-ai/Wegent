# SPDX-FileCopyrightText: 2026 Weibo, Inc.
#
# SPDX-License-Identifier: Apache-2.0

from __future__ import annotations

import io
import json
import re
import zipfile
from collections.abc import Callable
from datetime import datetime
from pathlib import Path
from unittest.mock import MagicMock

import pytest
from llama_index.core import Document
from llama_index.core.schema import MetadataMode, TextNode
from openpyxl import Workbook
from openpyxl.worksheet.formula import ArrayFormula, DataTableFormula

import knowledge_engine.excel as excel_module
from knowledge_engine.excel import (
    ExcelCell,
    ExcelRow,
    ExcelSheet,
    format_excel_retrieval_prefix,
    format_excel_row_compact,
    format_excel_row_readable,
    format_excel_sheet_header,
    normalize_excel_value,
    parse_excel_fragment_line,
    parse_excel_row_line,
    read_excel_sheets,
    serialize_excel_row,
    serialize_excel_rows,
    serialize_excel_sheet_compact,
)
from knowledge_engine.index.indexer import (
    DocumentIndexer,
    load_source_documents,
    sanitize_documents,
    sanitize_metadata,
)
from knowledge_engine.storage.chunk_metadata import ChunkMetadata


def _build_workbook(
    cells: dict[str, object],
    *,
    sheet_name: str = "Sheet1",
) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    for coordinate, value in cells.items():
        sheet[coordinate] = value
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _rewrite_sheet_xml(
    payload: bytes,
    transform: Callable[[bytes], bytes],
) -> bytes:
    output = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(payload), "r") as source:
        with zipfile.ZipFile(output, "w") as target:
            for item in source.infolist():
                item_payload = source.read(item)
                if item.filename == "xl/worksheets/sheet1.xml":
                    item_payload = transform(item_payload)
                target.writestr(item, item_payload)
    return output.getvalue()


def _replace_dimension(payload: bytes, declared_range: str) -> bytes:
    def transform(sheet_xml: bytes) -> bytes:
        rewritten, count = re.subn(
            rb'<dimension ref="[^"]+"/>',
            f'<dimension ref="{declared_range}"/>'.encode(),
            sheet_xml,
            count=1,
        )
        assert count == 1
        return rewritten

    return _rewrite_sheet_xml(payload, transform)


def _add_formula_cached_value(payload: bytes, cached_value: int) -> bytes:
    def transform(sheet_xml: bytes) -> bytes:
        rewritten, count = re.subn(
            rb'(<c r="A2"><f>1\+1</f><v>)(</v></c>)',
            rf"\g<1>{cached_value}\g<2>".encode(),
            sheet_xml,
            count=1,
        )
        assert count == 1
        return rewritten

    return _rewrite_sheet_xml(payload, transform)


@pytest.mark.parametrize("declared_range", ["A1", "A1:B2"])
def test_reader_ignores_incorrect_dimension(declared_range: str) -> None:
    payload = _replace_dimension(
        _build_workbook(
            {
                "A1": "first",
                "C2": "middle",
                "D3": "last",
            }
        ),
        declared_range,
    )

    [sheet] = read_excel_sheets(payload)

    assert sheet.rows == (
        ExcelRow(1, (ExcelCell(1, "first"),)),
        ExcelRow(2, (ExcelCell(3, "middle"),)),
        ExcelRow(3, (ExcelCell(4, "last"),)),
    )


def test_reader_preserves_sparse_columns_and_falsey_values() -> None:
    [sheet] = read_excel_sheets(
        _build_workbook(
            {
                "A1": "Alice",
                "C1": 30,
                "C2": "only-third-column",
                "A3": 0,
                "B3": False,
                "C3": " ",
            }
        )
    )

    assert sheet.rows == (
        ExcelRow(1, (ExcelCell(1, "Alice"), ExcelCell(3, 30))),
        ExcelRow(2, (ExcelCell(3, "only-third-column"),)),
        ExcelRow(
            3,
            (
                ExcelCell(1, 0),
                ExcelCell(2, False),
                ExcelCell(3, " "),
            ),
        ),
    )


def test_reader_preserves_first_row_without_header_inference() -> None:
    [sheet] = read_excel_sheets(
        _build_workbook(
            {
                "A1": "Alice",
                "B1": 30,
                "A2": "Bob",
                "B2": 40,
            }
        )
    )

    assert [serialize_excel_row(row) for row in sheet.rows] == [
        '{"source_row":1,"cells":[[1,"Alice"],[2,30]]}',
        '{"source_row":2,"cells":[[1,"Bob"],[2,40]]}',
    ]


def test_reader_loads_workbook_once_and_closes_it(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    workbook = MagicMock()
    workbook.worksheets = []
    load_workbook = MagicMock(return_value=workbook)
    monkeypatch.setattr(excel_module, "load_workbook", load_workbook)

    assert read_excel_sheets(b"xlsx") == ()
    load_workbook.assert_called_once()
    assert load_workbook.call_args.kwargs == {
        "read_only": True,
        "data_only": False,
    }
    workbook.close.assert_called_once_with()


def _corrupt_stylesheet(payload: bytes) -> bytes:
    """Rewrite xl/styles.xml so openpyxl raises the PR #1007 TypeError.

    A fills collection whose declared count exceeds its actual entries
    makes apply_stylesheet fill from an exhausted iterator, producing
    TypeError: expected <class 'openpyxl.styles.fills.Fill'>.
    """
    corrupt = (
        b'<?xml version="1.0"?>'
        b'<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        b'<fonts count="1"><font/></fonts><fills count="3"><fill/></fills>'
        b'<borders count="1"><border/></borders>'
        b'<cellStyleXfs count="1"><xf/></cellStyleXfs>'
        b'<cellXfs count="1"><xf fillId="0" fontId="0" borderId="0" xfId="0"/></cellXfs>'
        b"</styleSheet>"
    )
    output = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(payload), "r") as source:
        with zipfile.ZipFile(output, "w") as target:
            for item in source.infolist():
                item_payload = source.read(item)
                if item.filename == "xl/styles.xml":
                    item_payload = corrupt
                target.writestr(item, item_payload)
    return output.getvalue()


def test_reader_repairs_corrupted_stylesheet_without_global_patch() -> None:
    """A corrupted-stylesheet workbook (PR #1007) is repaired at the zip
    level and read through openpyxl's normal path.

    The retry replaces xl/styles.xml with a degenerate legal stylesheet
    instead of monkey-patching apply_stylesheet, so no global state is
    modified and concurrent normal loads can never observe a patched
    openpyxl.
    """
    import openpyxl.reader.excel as reader_module

    original_apply_stylesheet = reader_module.apply_stylesheet
    [sheet] = read_excel_sheets(
        _corrupt_stylesheet(_build_workbook({"A1": "Alice", "B1": 30}))
    )

    # Cell data survives the repair with faithful coordinates.
    assert sheet.name == "Sheet1"
    assert [serialize_excel_row(row) for row in sheet.rows] == [
        '{"source_row":1,"cells":[[1,"Alice"],[2,30]]}',
    ]
    # The global openpyxl entry point was never touched.
    assert reader_module.apply_stylesheet is original_apply_stylesheet


def test_reader_repairs_stylesheet_once_across_both_passes() -> None:
    """Formula workbooks load twice; the repair must not rebuild the zip
    for the second pass when the source is corrupt."""
    payload = _build_workbook({"A1": "Alice", "A2": "=1+1"})
    cached = _add_formula_cached_value(payload, cached_value=2)
    corrupt = _corrupt_stylesheet(cached)

    load_calls: list[bytes] = []
    real_load_workbook = excel_module.load_workbook

    def counting_load(source, **kwargs):
        load_calls.append(source.read() if hasattr(source, "read") else source)
        if hasattr(source, "seek"):
            source.seek(0)
        return real_load_workbook(source, **kwargs)

    with pytest.MonkeyPatch.context() as patches:
        patches.setattr(excel_module, "load_workbook", counting_load)
        [sheet] = read_excel_sheets(corrupt)

    # Both passes succeeded on repaired bytes, formula plus cached result.
    assert serialize_excel_row(sheet.rows[1]) == (
        '{"source_row":2,"cells":[[1,{"formula":"=1+1","value":2}]]}'
    )
    # One failed attempt on the original bytes, then one repair reused by
    # both passes: the two successful calls received byte-identical data.
    assert len(load_calls) == 3
    assert load_calls[1] == load_calls[2]
    assert load_calls[0] != load_calls[1]


def test_reader_propagates_unrelated_type_error_without_retry(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    load_workbook = MagicMock(side_effect=TypeError("invalid workbook input"))
    monkeypatch.setattr(excel_module, "load_workbook", load_workbook)

    with pytest.raises(TypeError, match="invalid workbook input"):
        read_excel_sheets(b"xlsx")

    load_workbook.assert_called_once()


def test_reader_preserves_uncached_formula_expression() -> None:
    [sheet] = read_excel_sheets(
        _build_workbook(
            {"A1": "Result", "A2": "=1+1"},
            sheet_name="Formula",
        )
    )

    assert serialize_excel_row(sheet.rows[1]) == (
        '{"source_row":2,"cells":[[1,"=1+1"]]}'
    )


def test_reader_merges_cached_result_into_formula_cell() -> None:
    payload = _add_formula_cached_value(
        _build_workbook(
            {"A1": "Result", "A2": "=1+1"},
            sheet_name="Formula",
        ),
        cached_value=2,
    )

    [sheet] = read_excel_sheets(payload)

    # A formula cell stores two facts; both are transcribed.
    line = serialize_excel_row(sheet.rows[1])
    assert line == '{"source_row":2,"cells":[[1,{"formula":"=1+1","value":2}]]}'
    assert parse_excel_row_line(line) == sheet.rows[1]


def test_reader_treats_equals_prefixed_string_as_data() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "=literal"
    sheet["A1"].data_type = "s"
    payload = io.BytesIO()
    workbook.save(payload)
    workbook.close()

    [parsed] = read_excel_sheets(payload.getvalue())

    assert serialize_excel_row(parsed.rows[0]) == (
        '{"source_row":1,"cells":[[1,"=literal"]]}'
    )


def test_reader_reads_cached_pass_only_when_formulas_exist(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    payload = _add_formula_cached_value(_build_workbook({"A2": "=1+1"}), 2)
    calls: list[dict] = []
    real_load_workbook = excel_module.load_workbook

    def recording_load_workbook(*args, **kwargs):  # noqa: ANN001, ANN002, ANN202
        calls.append(kwargs)
        return real_load_workbook(*args, **kwargs)

    monkeypatch.setattr(excel_module, "load_workbook", recording_load_workbook)

    [sheet] = read_excel_sheets(payload)

    assert [call["data_only"] for call in calls] == [False, True]
    assert sheet.rows[0].cells[0].value == {"formula": "=1+1", "value": 2}


def test_reader_serializes_array_formula_without_object_address() -> None:
    [sheet] = read_excel_sheets(
        _build_workbook(
            {"A1": ArrayFormula("A1:A2", "=ROW(A1:A2)")},
            sheet_name="Array Formula",
        )
    )

    assert serialize_excel_row(sheet.rows[0]) == (
        '{"source_row":1,"cells":[[1,{"formula_type":"array",'
        '"ref":"A1:A2","text":"=ROW(A1:A2)"}]]}'
    )


def test_reader_serializes_data_table_formula_without_object_address() -> None:
    [sheet] = read_excel_sheets(
        _build_workbook(
            {
                "A1": DataTableFormula(
                    "A1:A2",
                    ca=True,
                    dtr=True,
                    r1="B1",
                )
            },
            sheet_name="Data Table Formula",
        )
    )

    assert serialize_excel_row(sheet.rows[0]) == (
        '{"source_row":1,"cells":[[1,{"formula_type":"dataTable",'
        '"ref":"A1:A2","ca":true,"dt2D":false,"dtr":true,'
        '"r1":"B1","r2":null,"del1":false,"del2":false}]]}'
    )


def test_normalizer_retains_string_fallback_for_unknown_values() -> None:
    class UnknownExcelValue:
        def __str__(self) -> str:
            return "best-effort value"

    assert normalize_excel_value(UnknownExcelValue()) == "best-effort value"


def test_reader_normalizes_values_for_stable_serialization() -> None:
    [sheet] = read_excel_sheets(
        _build_workbook(
            {
                "A1": datetime(2026, 1, 2, 3, 4, 5),
                "B1": 2.0,
                "C1": "line 1\nline 2",
            }
        )
    )

    assert serialize_excel_row(sheet.rows[0]) == (
        '{"source_row":1,"cells":[[1,"2026-01-02T03:04:05"],'
        '[2,2],[3,"line 1\\nline 2"]]}'
    )


def test_source_reader_creates_one_document_per_non_empty_sheet(
    tmp_path: Path,
) -> None:
    workbook = Workbook()
    march = workbook.active
    march.title = "3月"
    march["A1"] = "March data"
    june = workbook.create_sheet("6月")
    june["C2"] = "June data"
    workbook.create_sheet("空表")
    path = tmp_path / "courses.xlsx"
    workbook.save(path)
    workbook.close()

    documents = load_source_documents(str(path), ".xlsx")

    assert [document.metadata["sheet_name"] for document in documents] == [
        "3月",
        "6月",
    ]
    assert documents[0].text == '{"source_row":1,"cells":[[1,"March data"]]}'
    assert documents[1].text == '{"source_row":2,"cells":[[3,"June data"]]}'
    assert sanitize_metadata(documents[1].metadata)["sheet_name"] == "6月"


def test_readable_row_keeps_coordinates_and_falsey_values() -> None:
    row = ExcelRow(
        3,
        (
            ExcelCell(1, 0),
            ExcelCell(2, False),
            ExcelCell(5, "张三"),
        ),
    )

    assert format_excel_row_readable(row) == (
        'Row 3, Column 1: 0; Column 2: false; Column 5: "张三"'
    )


def test_readable_row_inlines_structured_formula_values() -> None:
    row = ExcelRow(
        1,
        (
            ExcelCell(
                1,
                {"formula_type": "array", "ref": "A1:A2", "text": "=ROW(A1:A2)"},
            ),
        ),
    )

    assert format_excel_row_readable(row) == (
        'Row 1, Column 1: {"formula_type":"array","ref":"A1:A2",'
        '"text":"=ROW(A1:A2)"}'
    )


def test_readable_row_quotes_values_so_they_cannot_forge_structure() -> None:
    row = ExcelRow(
        5,
        (
            ExcelCell(1, "line1\nline2"),
            ExcelCell(2, "Alice; Column 9: forged"),
        ),
    )

    readable = format_excel_row_readable(row)

    assert "\n" not in readable
    assert readable == (
        'Row 5, Column 1: "line1\\nline2"; Column 2: "Alice; Column 9: forged"'
    )


def test_compact_row_keeps_dense_cells_bare_and_sparse_cells_marked() -> None:
    row = ExcelRow(
        2,
        (
            ExcelCell(1, "Alice"),
            ExcelCell(2, 100),
            ExcelCell(4, "x | y"),
        ),
    )

    assert format_excel_row_compact(row) == 'R2: "Alice" | 100 | [4]="x | y"'


def test_compact_row_marks_a_sparse_first_cell() -> None:
    row = ExcelRow(3, (ExcelCell(3, "备注"),))

    assert format_excel_row_compact(row) == 'R3: [3]="备注"'


def test_compact_row_quotes_strings_so_they_cannot_forge_structure() -> None:
    row = ExcelRow(
        5,
        (
            ExcelCell(1, "line1\nline2"),
            ExcelCell(2, "[9]=forged | R7: fake"),
        ),
    )

    compact = format_excel_row_compact(row)

    assert "\n" not in compact
    assert compact == 'R5: "line1\\nline2" | "[9]=forged | R7: fake"'


def test_compact_row_renders_merged_formula_with_its_cached_result() -> None:
    row = ExcelRow(2, (ExcelCell(1, {"formula": "=1+1", "value": 2}),))

    assert format_excel_row_compact(row) == 'R2: {"formula":"=1+1","value":2}'


def test_compact_sheet_serializes_name_and_rows() -> None:
    sheet = ExcelSheet(
        name="数据",
        rows=(ExcelRow(1, (ExcelCell(1, "a"), ExcelCell(2, False))),),
    )

    assert (
        serialize_excel_sheet_compact(sheet) == '--- Sheet: 数据 ---\nR1: "a" | false'
    )
    empty = ExcelSheet(name="空", rows=())
    assert serialize_excel_sheet_compact(empty) == "--- Sheet: 空 ---"


def test_sheet_header_collapses_newlines_so_names_cannot_forge_rows() -> None:
    assert format_excel_sheet_header("数据\nR7: fake | 100") == (
        "--- Sheet: 数据 R7: fake | 100 ---"
    )
    # The compact projection and the indexing display prefix share the header.
    sheet = ExcelSheet(
        name="S\nR7: fake | 100",
        rows=(ExcelRow(1, (ExcelCell(1, "value"),)),),
    )
    text = serialize_excel_sheet_compact(sheet)
    assert text.split("\n") == ["--- Sheet: S R7: fake | 100 ---", 'R1: "value"']


def test_serialized_rows_round_trip_line_by_line() -> None:
    [sheet] = read_excel_sheets(_build_workbook({"A1": "Alice", "C1": 30, "A2": "Bob"}))

    rows = [
        parse_excel_row_line(line)
        for line in serialize_excel_rows(sheet.rows).split("\n")
    ]

    assert rows == list(sheet.rows)


@pytest.mark.parametrize(
    "line",
    [
        '{"source_row":2,"cells":[[1,"Bo',  # chunk boundary cut the row
        "not json at all",
        '{"other":1}',
        '{"source_row":true,"cells":[[1,"x"]]}',  # bool is not an int coordinate
        '{"source_row":1,"cells":[[1,NaN]]}',  # non-canonical float token
        '{"source_row":1,"cells":[]}',  # canonical rows always have cells
        '{"source_row":1,"cells":[[1,null]]}',  # populated cells are never null
        '{"source_row": 1, "cells": [[1, 30]]}',  # non-canonical spacing
    ],
)
def test_parse_row_line_rejects_non_canonical_lines(line: str) -> None:
    with pytest.raises(ValueError, match="canonical"):
        parse_excel_row_line(line)


def test_retrieval_prefix_renders_worksheet_line() -> None:
    assert format_excel_retrieval_prefix("FAQ") == "Worksheet: FAQ\n"


@pytest.mark.parametrize("empty_name", ["", "\n", "  \n "])
def test_retrieval_prefix_empty_name_emits_nothing(empty_name: str) -> None:
    assert format_excel_retrieval_prefix(empty_name) == ""


def test_retrieval_prefix_collapses_newlines_so_names_cannot_forge_rows() -> None:
    assert format_excel_retrieval_prefix("数据\nRow 1: forged") == (
        "Worksheet: 数据 Row 1: forged\n"
    )


def test_retrieval_prefix_caps_overlong_names_with_marker() -> None:
    assert format_excel_retrieval_prefix("x" * 100) == f"Worksheet: {'x' * 47}…\n"


def test_excel_index_preserves_sheet_context_on_every_chunk() -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for name in ("3月", "6月"):
        sheet = workbook.create_sheet(name)
        for index in range(12):
            sheet.append([index, f"{name}课程{index}" + "内容" * 20])
    payload = io.BytesIO()
    workbook.save(payload)
    workbook.close()

    storage_backend = MagicMock()
    storage_backend.index_with_metadata.return_value = {"indexed_count": 1}
    indexer = DocumentIndexer(
        storage_backend=storage_backend,
        embed_model=MagicMock(),
        splitter_config={
            "chunk_strategy": "flat",
            "format_enhancement": "file_aware",
            "flat_config": {
                "chunk_size": 180,
                "chunk_overlap": 0,
                "separator": "\n",
            },
        },
        file_extension=".xlsx",
    )

    result = indexer.index_from_binary(
        binary_data=payload.getvalue(),
        file_extension=".xlsx",
        chunk_metadata=ChunkMetadata(
            knowledge_id="1",
            doc_ref="courses",
            source_file="courses.xlsx",
            created_at="2026-08-14T00:00:00+00:00",
        ),
    )

    nodes = storage_backend.index_with_metadata.call_args.kwargs["nodes"]
    assert {node.metadata["sheet_name"] for node in nodes} == {"3月", "6月"}
    for node in nodes:
        expected_body = "\n".join(
            format_excel_row_readable(parse_excel_row_line(line))
            for line in node.text.split("\n")
        )
        expected_retrieval = (
            f"Worksheet: {node.metadata['sheet_name']}\n" + expected_body
        )
        # Row-atomic chunking keeps every canonical line intact, and the
        # worksheet prefix is paid out of the chunk budget, not on top.
        assert len(node.text) <= 180
        assert node.metadata["retrieval_text"] == expected_retrieval
        assert len(node.metadata["retrieval_text"]) <= 180
        assert node.metadata["display_text"] == (
            f"--- Sheet: {node.metadata['sheet_name']} ---\n" + node.text
        )
        prepared_node = node.model_copy(
            update={"text": node.metadata["retrieval_text"]}
        )
        embedding_content = prepared_node.get_content(metadata_mode=MetadataMode.EMBED)
        assert embedding_content.count(expected_retrieval) == 1
        assert "Row " in embedding_content
        assert node.metadata["display_text"] not in embedding_content
    assert all(
        item["content"].startswith("--- Sheet: ")
        for item in result["chunks_data"]["items"]
    )


def _index_workbook(
    splitter_config: dict,
    *,
    cells_per_row: int = 2,
    row_count: int = 12,
    long_cell: int = 0,
) -> tuple[list, dict, MagicMock]:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "数据"
    for row_index in range(row_count):
        for column_index in range(cells_per_row):
            value = f"r{row_index}c{column_index}" + "内容" * 10
            if long_cell and row_index == 3 and column_index == 0:
                value = "长" * long_cell
            sheet.cell(row=row_index + 1, column=column_index + 1, value=value)
    payload = io.BytesIO()
    workbook.save(payload)
    workbook.close()

    storage_backend = MagicMock()
    storage_backend.index_with_metadata.return_value = {"indexed_count": 1}
    indexer = DocumentIndexer(
        storage_backend=storage_backend,
        embed_model=MagicMock(),
        splitter_config=splitter_config,
        file_extension=".xlsx",
    )
    result = indexer.index_from_binary(
        binary_data=payload.getvalue(),
        file_extension=".xlsx",
        chunk_metadata=ChunkMetadata(
            knowledge_id="1",
            doc_ref="data",
            source_file="data.xlsx",
            created_at="2026-08-17T00:00:00+00:00",
        ),
    )
    nodes = storage_backend.index_with_metadata.call_args.kwargs["nodes"]
    return nodes, result, storage_backend


def test_excel_cells_with_data_urls_survive_the_sanitizer_stage() -> None:
    """Data-URL cell values must not be rewritten before the strict parser."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "导出"
    sheet["A1"] = "image: data:image/png;base64,iVBORw0KGgoAAAANSUhEUg"
    payload = io.BytesIO()
    workbook.save(payload)
    workbook.close()

    storage_backend = MagicMock()
    storage_backend.index_with_metadata.return_value = {"indexed_count": 1}
    indexer = DocumentIndexer(
        storage_backend=storage_backend,
        embed_model=MagicMock(),
        splitter_config={"chunk_strategy": "flat", "format_enhancement": "none"},
        file_extension=".xlsx",
    )

    indexer.index_from_binary(
        binary_data=payload.getvalue(),
        file_extension=".xlsx",
        chunk_metadata=ChunkMetadata(
            knowledge_id="1",
            doc_ref="export",
            source_file="export.xlsx",
            created_at="2026-08-17T00:00:00+00:00",
        ),
    )

    nodes = storage_backend.index_with_metadata.call_args.kwargs["nodes"]
    assert any("iVBORw0KGgoAAAANSUhEUg" in node.text for node in nodes)


def test_excel_documents_with_empty_sheet_name_skip_the_sanitizer() -> None:
    """An empty sheet name is still an Excel document; presence is the signal."""
    data_url = "image: data:image/png;base64,iVBORw0KGgoAAAANSUhEUg"
    document = Document(
        text=f'{{"source_row":1,"cells":[[1,{json.dumps(data_url)}]]}}',
        metadata={"sheet_name": ""},
    )

    documents = sanitize_documents([document], sanitize_inline_images=True)

    # The canonical row text survived verbatim; the sanitizer never ran.
    assert documents[0].text == (
        '{"source_row":1,"cells":[[1,'
        '"image: data:image/png;base64,iVBORw0KGgoAAAANSUhEUg"]]}'
    )
    assert documents[0].metadata["sheet_name"] == ""


def test_index_document_resolves_extension_per_call(tmp_path: Path) -> None:
    """A reused indexer never leaks file_extension state across files."""
    path = tmp_path / "book.xlsx"
    workbook = Workbook()
    workbook.active["A1"] = "value"
    workbook.save(path)
    workbook.close()

    storage_backend = MagicMock()
    storage_backend.index_with_metadata.return_value = {"indexed_count": 1}
    indexer = DocumentIndexer(
        storage_backend=storage_backend,
        embed_model=MagicMock(),
        splitter_config={"chunk_strategy": "flat", "format_enhancement": "none"},
    )

    indexer.index_document(
        file_path=str(path),
        chunk_metadata=ChunkMetadata(
            knowledge_id="1",
            doc_ref="book",
            source_file="book.xlsx",
            created_at="2026-08-17T00:00:00+00:00",
        ),
    )

    assert indexer.file_extension is None
    nodes = storage_backend.index_with_metadata.call_args.kwargs["nodes"]
    assert all(node.metadata.get("retrieval_text") for node in nodes)


def test_excel_semantic_strategy_degrades_to_row_atomic_chunks() -> None:
    nodes, result, _ = _index_workbook(
        {"chunk_strategy": "semantic", "format_enhancement": "none"},
        long_cell=2000,
    )

    assert result["chunks_data"]["splitter_subtype"] == "excel_rows"
    for node in nodes:
        for line in node.text.split("\n"):
            # No chunk is a cut JSON fragment, even for the 2000-char cell.
            if '"fragment_index"' in line:
                parse_excel_fragment_line(line)
            else:
                parse_excel_row_line(line)
        assert node.metadata["retrieval_text"]


def test_excel_hierarchical_strategy_builds_row_atomic_parents() -> None:
    nodes, result, storage_backend = _index_workbook(
        {
            "chunk_strategy": "hierarchical",
            "format_enhancement": "none",
            "hierarchical_config": {
                "parent_chunk_size": 512,
                "child_chunk_size": 160,
                "child_chunk_overlap": 0,
            },
        },
        long_cell=1500,
    )

    parent_nodes = storage_backend.save_parent_nodes.call_args.kwargs["parent_nodes"]
    assert parent_nodes
    assert len(nodes) > len(parent_nodes)
    for parent in parent_nodes:
        for line in parent.text.split("\n"):
            if '"fragment_index"' in line:
                parse_excel_fragment_line(line)
            else:
                parse_excel_row_line(line)
        assert parent.metadata["retrieval_text"]
        assert len(parent.text) <= 512
    for child in nodes:
        assert len(child.text) <= 160
        assert child.metadata["parent_node_id"] in {
            parent.node_id for parent in parent_nodes
        }
    assert result["chunks_data"]["splitter_subtype"] == "excel_rows"
