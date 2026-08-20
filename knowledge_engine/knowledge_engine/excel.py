# SPDX-FileCopyrightText: 2026 Weibo, Inc.
#
# SPDX-License-Identifier: Apache-2.0

"""Shared Excel loading and lossless row serialization."""

from __future__ import annotations

import io
import json
import logging
import math
import zipfile
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Any, BinaryIO, TypeAlias

from openpyxl import load_workbook
from openpyxl.worksheet.formula import ArrayFormula, DataTableFormula

logger = logging.getLogger(__name__)

ExcelScalar: TypeAlias = str | int | float | bool
# Structured values may nest: a formula wrapper can carry its cached result.
ExcelStructuredValue: TypeAlias = dict[str, "ExcelValue | None"]
ExcelValue: TypeAlias = ExcelScalar | ExcelStructuredValue
ExcelSource: TypeAlias = str | Path | bytes | BinaryIO

EXCEL_SOURCE_EXTENSIONS: frozenset[str] = frozenset({".xlsx"})

# Hard cap for the sheet name inside retrieval/display prefixes, in
# characters. XLSX caps worksheet names at 31 characters, so this only
# guards hand-crafted files while keeping the prefix bounded for the
# budget. Indexing may shrink the cap further when a minimal chunk budget
# must reserve room for cell fragments (see excel_rows.py).
MAX_EXCEL_CONTEXT_NAME = 48


@dataclass(frozen=True)
class ExcelCell:
    """A populated cell at its 1-based physical source column."""

    source_column: int
    value: ExcelValue


@dataclass(frozen=True)
class ExcelRow:
    """A populated row at its 1-based physical source row."""

    source_row: int
    cells: tuple[ExcelCell, ...]


@dataclass(frozen=True)
class ExcelSheet:
    """A worksheet and its populated physical rows."""

    name: str
    rows: tuple[ExcelRow, ...]


@dataclass(frozen=True)
class ExcelCellFragment:
    """A sized fragment of one overlong cell value.

    ``value`` holds a substring of the original string value
    (``encoding="text"``) or of its canonical JSON encoding
    (``encoding="json"``, for structured values such as formulas).
    """

    source_row: int
    source_column: int
    fragment_index: int
    fragment_count: int
    value: str
    encoding: str = "text"


def _normalize_formula_flag(value: Any) -> bool:
    """Normalize an OOXML boolean flag loaded as a bool or string."""
    if isinstance(value, str):
        return value.lower() in {"1", "true"}
    return bool(value)


def _normalize_formula_value(
    value: ArrayFormula | DataTableFormula,
) -> ExcelStructuredValue:
    """Normalize an OpenPyXL formula wrapper to stable JSON data."""
    if isinstance(value, ArrayFormula):
        return {
            "formula_type": "array",
            "ref": value.ref,
            "text": value.text,
        }
    return {
        "formula_type": "dataTable",
        "ref": value.ref,
        "ca": _normalize_formula_flag(value.ca),
        "dt2D": _normalize_formula_flag(value.dt2D),
        "dtr": _normalize_formula_flag(value.dtr),
        "r1": value.r1,
        "r2": value.r2,
        "del1": _normalize_formula_flag(value.del1),
        "del2": _normalize_formula_flag(value.del2),
    }


def normalize_excel_value(value: Any) -> ExcelValue:
    """Normalize an OpenPyXL value to deterministic JSON data."""
    if isinstance(value, (ArrayFormula, DataTableFormula)):
        return _normalize_formula_value(value)
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, (date, time)):
        return value.isoformat()
    if isinstance(value, timedelta):
        return str(value)
    if isinstance(value, float) and not math.isfinite(value):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return int(value)
    if isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def _is_formula_cell(cell: Any) -> bool:
    """Detect formula cells by type, not by value prefix.

    A literal string starting with "=" is data, not a formula, so the
    cell's data_type (or a formula wrapper object) is the only reliable
    signal.
    """
    return cell.data_type == "f" or isinstance(
        cell.value, (ArrayFormula, DataTableFormula)
    )


def _read_sheet(worksheet: Any) -> tuple[ExcelSheet, frozenset[tuple[int, int]]]:
    """Read populated cells without trusting the worksheet dimension.

    Returns the sheet plus the physical coordinates of its formula cells,
    so the caller can merge cached results from a second data_only pass.
    """
    worksheet.reset_dimensions()
    rows: list[ExcelRow] = []
    formula_coords: set[tuple[int, int]] = set()
    for source_cells in worksheet.iter_rows():
        populated_cells = [cell for cell in source_cells if cell.value is not None]
        if not populated_cells:
            continue
        formula_coords.update(
            (cell.row, cell.column)
            for cell in populated_cells
            if _is_formula_cell(cell)
        )
        rows.append(
            ExcelRow(
                source_row=populated_cells[0].row,
                cells=tuple(
                    ExcelCell(
                        source_column=cell.column,
                        value=normalize_excel_value(cell.value),
                    )
                    for cell in populated_cells
                ),
            )
        )
    return ExcelSheet(name=worksheet.title, rows=tuple(rows)), frozenset(formula_coords)


def _materialize_bytes(source: ExcelSource) -> bytes:
    """Read the workbook source into bytes once, for repeated loads."""
    if isinstance(source, bytes):
        return source
    if isinstance(source, Path):
        return source.read_bytes()
    if isinstance(source, str):
        return Path(source).read_bytes()
    return source.read()


def _open_workbook(source: Any, *, data_only: bool) -> Any:
    """Load a workbook; read_only keeps coordinates faithful to the file."""
    return load_workbook(source, read_only=True, data_only=data_only)


def _is_stylesheet_type_error(exc: TypeError) -> bool:
    """Match only the verified OpenPyXL stylesheet parsing failure."""
    return "openpyxl.styles" in str(exc)


# A legal stylesheet with no style definitions at all: openpyxl accepts it,
# every style lookup degrades to the "General" default, and cell data is
# untouched. This is byte-for-byte what skipping stylesheet parsing used to
# produce, so the repair changes no observable output for damaged files.
_DEGENERATE_STYLESHEET = (
    b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
    b'<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
    b'<numFmts count="0"/><fonts count="0"/><fills count="0"/><borders count="0"/>'
    b'<cellStyleXfs count="0"/><cellXfs count="0"/><cellStyles count="0"/>'
    b'<dxfs count="0"/><tableStyles count="0"/>'
    b"</styleSheet>"
)


def _repair_stylesheet(data: bytes) -> bytes:
    """Replace xl/styles.xml with a degenerate legal stylesheet.

    Styles carry no cell data, so this only drops the already-unreadable
    style definitions of a damaged workbook; every other zip entry is
    copied through unchanged. openpyxl then loads the repaired bytes
    through its normal path — no global state is modified.
    """
    output = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(data), "r") as source:
        with zipfile.ZipFile(output, "w") as target:
            for item in source.infolist():
                payload = source.read(item)
                if item.filename == "xl/styles.xml":
                    payload = _DEGENERATE_STYLESHEET
                target.writestr(item, payload)
    return output.getvalue()


def read_excel_sheets(source: ExcelSource) -> tuple[ExcelSheet, ...]:
    """Read an XLSX, preserving formulas, cached results, and coordinates.

    A formula cell physically stores two facts — the expression and the
    last calculated result — so both are transcribed. The formula pass
    (data_only=False) is authoritative for coordinates; a second pass
    (data_only=True) merges cached results into formula cells as
    ``{"formula": ..., "value": ...}``. Formula cells that were never
    calculated keep the plain expression. The second pass is skipped
    entirely when no formula cells exist.

    A workbook whose stylesheet cannot be parsed (PR #1007 files) is
    repaired once — its xl/styles.xml replaced with a degenerate legal
    one — and both passes load the repaired bytes; openpyxl itself is
    never patched.
    """
    data = _materialize_bytes(source)
    try:
        formula_book = _open_workbook(io.BytesIO(data), data_only=False)
    except TypeError as exc:
        if not _is_stylesheet_type_error(exc):
            raise
        logger.warning(
            "Retrying workbook load with a repaired stylesheet after: %s", exc
        )
        data = _repair_stylesheet(data)
        formula_book = _open_workbook(io.BytesIO(data), data_only=False)
    try:
        sheets = tuple(_read_sheet(worksheet) for worksheet in formula_book.worksheets)
    finally:
        formula_book.close()
    if not any(formula_coords for _, formula_coords in sheets):
        return tuple(sheet for sheet, _ in sheets)
    cached_book = _open_workbook(io.BytesIO(data), data_only=True)
    try:
        merged = []
        for (sheet, formula_coords), cached_worksheet in zip(
            sheets, cached_book.worksheets
        ):
            if not formula_coords:
                merged.append(sheet)
                continue
            cached = _read_cached_values(cached_worksheet, formula_coords)
            merged.append(_merge_cached_values(sheet, cached))
        return tuple(merged)
    finally:
        cached_book.close()


def _read_cached_values(
    worksheet: Any, coords: frozenset[tuple[int, int]]
) -> dict[tuple[int, int], Any]:
    """Read cached results for formula coordinates from a data_only sheet."""
    worksheet.reset_dimensions()
    values: dict[tuple[int, int], Any] = {}
    for source_cells in worksheet.iter_rows():
        for cell in source_cells:
            if cell.value is not None and (cell.row, cell.column) in coords:
                values[(cell.row, cell.column)] = cell.value
    return values


def _merge_cached_values(
    sheet: ExcelSheet, cached: dict[tuple[int, int], Any]
) -> ExcelSheet:
    """Wrap each formula cell with its cached result when one exists."""
    rows = []
    for row in sheet.rows:
        cells = tuple(
            _merge_cached_value(row.source_row, cell, cached) for cell in row.cells
        )
        rows.append(ExcelRow(source_row=row.source_row, cells=cells))
    return ExcelSheet(name=sheet.name, rows=tuple(rows))


def _merge_cached_value(
    source_row: int, cell: ExcelCell, cached: dict[tuple[int, int], Any]
) -> ExcelCell:
    key = (source_row, cell.source_column)
    if key not in cached:
        return cell
    merged: ExcelValue = {
        "formula": cell.value,
        "value": normalize_excel_value(cached[key]),
    }
    return ExcelCell(source_column=cell.source_column, value=merged)


def serialize_excel_row(row: ExcelRow) -> str:
    """Serialize a sparse row without losing its physical coordinates."""
    return json.dumps(
        {
            "source_row": row.source_row,
            "cells": [[cell.source_column, cell.value] for cell in row.cells],
        },
        ensure_ascii=False,
        separators=(",", ":"),
        allow_nan=False,
    )


def serialize_excel_rows(rows: tuple[ExcelRow, ...]) -> str:
    """Serialize physical rows without worksheet presentation context."""
    return "\n".join(serialize_excel_row(row) for row in rows)


# Attachment projection: compact row text for LLM context. It derives from
# the same ExcelRow model as the canonical serialization, but serves
# consumers that read rather than parse: rows carry their number as an
# "R<n>:" prefix, dense cells join with " | ", sparse cells carry their
# column as "[<column>]=<value>", and strings stay JSON-quoted so values
# can never forge row structure. Round-trip parsing is intentionally not a
# property of this projection; the canonical serialization owns that.


def format_excel_row_compact(row: ExcelRow) -> str:
    """Render a row as compact attachment text for LLM context."""
    parts: list[str] = []
    expected_column = 1
    for cell in row.cells:
        value = _format_compact_value(cell.value)
        if cell.source_column == expected_column:
            parts.append(value)
        else:
            parts.append(f"[{cell.source_column}]={value}")
        expected_column = cell.source_column + 1
    return f"R{row.source_row}: " + " | ".join(parts)


def _format_compact_value(value: ExcelValue) -> str:
    """Render a value compactly: strings stay quoted, scalars stay bare."""
    if isinstance(value, str):
        return json.dumps(value, ensure_ascii=False)
    return dumps_excel_value(value)


def _single_line_value(value: str) -> str:
    """Collapse newlines so a name cannot forge line structure."""
    return " ".join(value.splitlines())


def format_excel_sheet_header(
    name: str, *, name_cap: int = MAX_EXCEL_CONTEXT_NAME
) -> str:
    """Render the sheet header from a newline-collapsed name, or "" if empty.

    Mirrors the retrieval prefix: an empty name emits no header at all so a
    hand-crafted empty one cannot cost every chunk budget, and overlong
    names truncate to the cap.
    """
    collapsed = _truncate_sheet_name(name, name_cap)
    if not collapsed:
        return ""
    return f"--- Sheet: {collapsed} ---"


def format_excel_retrieval_prefix(
    sheet_name: str, *, name_cap: int = MAX_EXCEL_CONTEXT_NAME
) -> str:
    """Render the one-line retrieval prefix within the name cap.

    An empty (or newline-only) name emits no prefix at all: real workbooks
    always carry a title, and a hand-crafted empty one must not cost every
    chunk budget for a context line that carries no signal.
    """
    collapsed = _truncate_sheet_name(sheet_name, name_cap)
    if not collapsed:
        return ""
    return f"Worksheet: {collapsed}\n"


def _truncate_sheet_name(name: str, name_cap: int) -> str:
    """Collapse newlines and truncate to the cap, returning the result."""
    collapsed = _single_line_value(name).strip()
    if name_cap < 1 or not collapsed:
        return ""
    if len(collapsed) > name_cap:
        return collapsed[: name_cap - 1] + "…"
    return collapsed


def serialize_excel_sheet_compact(sheet: ExcelSheet) -> str:
    """Serialize a worksheet as compact attachment text with its name."""
    header = format_excel_sheet_header(sheet.name)
    if not header:
        return "\n".join(format_excel_row_compact(row) for row in sheet.rows)
    rows_text = "\n".join(format_excel_row_compact(row) for row in sheet.rows)
    return f"{header}\n{rows_text}" if rows_text else header


def dumps_excel_value(value: ExcelValue) -> str:
    """Encode a value with the one canonical JSON encoding."""
    return json.dumps(value, ensure_ascii=False, separators=(",", ":"), allow_nan=False)


def serialize_excel_fragment(fragment: ExcelCellFragment) -> str:
    """Serialize a cell fragment with its full coordinates."""
    return json.dumps(
        {
            "source_row": fragment.source_row,
            "source_column": fragment.source_column,
            "fragment_index": fragment.fragment_index,
            "fragment_count": fragment.fragment_count,
            "value": fragment.value,
            "encoding": fragment.encoding,
        },
        ensure_ascii=False,
        separators=(",", ":"),
        allow_nan=False,
    )


# Readable retrieval protocol: one line per row or fragment. Row lines are
# "Row <n>, Column <c>: <value>; ..." and fragment lines are
# "Row <n>, Column <c>, fragment <i> of <count>: <value>". Every value is
# JSON-encoded, so value text can never be confused with line structure.


def format_excel_row_readable(row: ExcelRow) -> str:
    """Render a row as readable text that keeps its physical coordinates."""
    cells = "; ".join(
        f"Column {cell.source_column}: {_format_readable_value(cell.value)}"
        for cell in row.cells
    )
    return f"Row {row.source_row}, {cells}"


def format_excel_fragment_readable(fragment: ExcelCellFragment) -> str:
    """Render a cell fragment as readable text with full coordinates."""
    value = json.dumps(fragment.value, ensure_ascii=False)
    return (
        f"Row {fragment.source_row}, Column {fragment.source_column}, "
        f"fragment {fragment.fragment_index} of {fragment.fragment_count}: "
        f"{value}"
    )


def _format_readable_value(value: ExcelValue) -> str:
    """Render a normalized value as JSON-encoded single-line text."""
    return dumps_excel_value(value)


def parse_excel_row_line(line: str) -> ExcelRow:
    """Parse one canonical serialized row line, rejecting anything else."""
    payload = _load_canonical_line(line, "row")
    source_row = payload.get("source_row")
    cells = payload.get("cells")
    # bool is a subclass of int, so check the exact type for coordinates.
    if (
        type(source_row) is not int
        or source_row < 1
        or not isinstance(cells, list)
        or not cells
    ):
        raise ValueError(f"not a canonical Excel row line: {line[:80]!r}")
    if any(
        not isinstance(cell, list)
        or len(cell) != 2
        or type(cell[0]) is not int
        or cell[0] < 1
        or cell[1] is None
        for cell in cells
    ):
        raise ValueError(f"not a canonical Excel row line: {line[:80]!r}")
    row = ExcelRow(
        source_row=source_row,
        cells=tuple(ExcelCell(source_column=cell[0], value=cell[1]) for cell in cells),
    )
    try:
        canonical = serialize_excel_row(row)
    except ValueError as exc:
        # Non-finite floats (NaN/Infinity) cannot be canonical output.
        raise ValueError(f"not a canonical Excel row line: {line[:80]!r}") from exc
    _assert_round_trip(line, canonical)
    return row


def parse_excel_fragment_line(line: str) -> ExcelCellFragment:
    """Parse one canonical serialized fragment line, rejecting anything else."""
    payload = _load_canonical_line(line, "fragment")
    fragment = ExcelCellFragment(
        source_row=payload.get("source_row"),
        source_column=payload.get("source_column"),
        fragment_index=payload.get("fragment_index"),
        fragment_count=payload.get("fragment_count"),
        value=payload.get("value"),
        encoding=payload.get("encoding") or "",
    )
    coordinates = (
        fragment.source_row,
        fragment.source_column,
        fragment.fragment_index,
        fragment.fragment_count,
    )
    if (
        any(type(value) is not int for value in coordinates)
        or fragment.source_row < 1
        or fragment.source_column < 1
        or not 1 <= fragment.fragment_index <= fragment.fragment_count
        or not isinstance(fragment.value, str)
        or fragment.encoding not in {"text", "json"}
    ):
        raise ValueError(f"not a canonical Excel fragment line: {line[:80]!r}")
    _assert_round_trip(line, serialize_excel_fragment(fragment))
    return fragment


def _load_canonical_line(line: str, kind: str) -> Any:
    """Load one JSON line, rejecting anything but an object payload."""
    try:
        payload = json.loads(line)
    except ValueError as exc:
        raise ValueError(f"not a canonical Excel {kind} line: {line[:80]!r}") from exc
    if not isinstance(payload, dict):
        raise ValueError(f"not a canonical Excel {kind} line: {line[:80]!r}")
    return payload


def _assert_round_trip(line: str, canonical: str) -> None:
    """Reject lines that are not exact canonical serializer output."""
    if canonical != line:
        raise ValueError(f"not canonical serializer output: {line[:80]!r}")
