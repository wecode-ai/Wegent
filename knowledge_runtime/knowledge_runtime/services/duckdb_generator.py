# SPDX-FileCopyrightText: 2026 Weibo, Inc.
#
# SPDX-License-Identifier: Apache-2.0

"""DuckDB file generation service for Excel/CSV data analysis.

Generates .duckdb files from Excel/CSV binary data fetched via ContentRef,
including automatic table detection, SUMMARIZE analysis, and sample data extraction.
"""

from __future__ import annotations

import hashlib
import logging
import os
import re
import shutil
import time
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import duckdb

from knowledge_runtime.config import get_settings
from knowledge_runtime.services.content_fetcher import ContentFetcher, ContentFetchError
from shared.models.data_analysis_protocol import DuckDBColumnInfo, DuckDBTableInfo
from shared.models.knowledge_runtime_protocol import ContentRef

logger = logging.getLogger(__name__)


@dataclass
class DuckDBGenerateResult:
    """Result of DuckDB file generation."""

    duckdb_bytes: bytes
    summary: dict[str, Any]  # SUMMARIZE output per table
    tables: list[DuckDBTableInfo] = field(default_factory=list)
    sample_data: dict[str, list[Any]] = field(default_factory=dict)
    generation_time_ms: float = 0.0
    source_file_hash: str = ""
    source_file_size: int = 0


class DuckDBGenerator:
    """Generates .duckdb files from Excel/CSV binary data.

    Handles multi-sheet Excel files, CSV/TSV imports, and legacy .xls format.
    Produces a DuckDB database file with SUMMARIZE metadata and sample data.
    """

    # Supported file extensions
    SUPPORTED_EXTENSIONS = {".xlsx", ".xls", ".csv", ".tsv"}

    def __init__(self) -> None:
        self._settings = get_settings()
        self._content_fetcher = ContentFetcher()

    async def generate(
        self,
        content_ref: ContentRef,
        source_file: str,
        file_extension: str,
    ) -> DuckDBGenerateResult:
        """Generate a .duckdb file from Excel/CSV data.

        Steps:
        1. Fetch binary data via ContentFetcher
        2. Check file size against DUCKDB_MAX_FILE_SIZE_MB
        3. Check available memory against DUCKDB_MIN_FREE_MEMORY_MB
        4. Import data into DuckDB write connection
        5. Run SUMMARIZE on all tables
        6. Extract sample data (first N rows)
        7. CHECKPOINT to merge WAL
        8. Close write connection
        9. Return duckdb bytes + summary + tables info

        Args:
            content_ref: Content reference for fetching the source file.
            source_file: Original filename of the source file.
            file_extension: File extension including dot (e.g., ".xlsx").

        Returns:
            DuckDBGenerateResult with the generated database and metadata.

        Raises:
            ValueError: If the file format is unsupported or file is too large.
            ContentFetchError: If content fetching fails.
            RuntimeError: If DuckDB generation fails.
        """
        start_time = time.monotonic()

        # Normalize extension
        ext = file_extension.lower().strip()
        if not ext.startswith("."):
            ext = "." + ext

        if ext not in self.SUPPORTED_EXTENSIONS:
            raise ValueError(
                f"Unsupported file extension: {ext}. "
                f"Supported: {sorted(self.SUPPORTED_EXTENSIONS)}"
            )

        # Step 1: Fetch binary data
        binary_data, resolved_source, resolved_ext = await self._content_fetcher.fetch(
            content_ref
        )

        # Override with caller-provided metadata
        if source_file:
            resolved_source = source_file
        if ext:
            resolved_ext = ext

        # Compute SHA256 hash of the SOURCE file (not the generated DuckDB)
        # This allows detecting whether the source file has changed on re-index.
        source_file_hash = hashlib.sha256(binary_data).hexdigest()
        source_file_size = len(binary_data)

        # Step 2: Check file size
        max_size_bytes = self._settings.duckdb_max_file_size_mb * 1024 * 1024
        if len(binary_data) > max_size_bytes:
            raise ValueError(
                f"File size ({len(binary_data) / 1024 / 1024:.1f} MB) exceeds "
                f"maximum allowed size ({self._settings.duckdb_max_file_size_mb} MB)"
            )

        # Step 3: Check available memory
        self._check_available_memory()

        # Step 4-8: Run DuckDB generation in thread (CPU-bound)
        result = await self._run_generation(
            binary_data=binary_data,
            source_file=resolved_source,
            file_extension=resolved_ext,
        )

        result.generation_time_ms = (time.monotonic() - start_time) * 1000
        result.source_file_hash = source_file_hash
        result.source_file_size = source_file_size
        return result

    def _check_available_memory(self) -> None:
        """Check that sufficient memory is available for DuckDB operations.

        Raises:
            RuntimeError: If insufficient memory is available.
        """
        try:
            import psutil

            available_mb = psutil.virtual_memory().available / (1024 * 1024)
            min_required_mb = self._settings.duckdb_min_free_memory_mb
            if available_mb < min_required_mb:
                raise RuntimeError(
                    f"Insufficient memory: {available_mb:.0f} MB available, "
                    f"{min_required_mb} MB required"
                )
        except ImportError:
            # psutil not available, skip memory check
            logger.warning("psutil not available, skipping memory check")

    async def _run_generation(
        self,
        binary_data: bytes,
        source_file: str,
        file_extension: str,
    ) -> DuckDBGenerateResult:
        """Run the CPU-bound DuckDB generation in a thread.

        Args:
            binary_data: Raw file content.
            source_file: Original filename.
            file_extension: Normalized file extension.

        Returns:
            DuckDBGenerateResult with the generated database and metadata.
        """
        import asyncio

        return await asyncio.to_thread(
            self._generate_sync,
            binary_data,
            source_file,
            file_extension,
        )

    def _generate_sync(
        self,
        binary_data: bytes,
        source_file: str,
        file_extension: str,
    ) -> DuckDBGenerateResult:
        """Synchronous DuckDB generation logic.

        Args:
            binary_data: Raw file content.
            source_file: Original filename.
            file_extension: Normalized file extension.

        Returns:
            DuckDBGenerateResult with the generated database and metadata.

        Raises:
            RuntimeError: If DuckDB generation fails.
        """
        import tempfile

        settings = self._settings
        temp_dir = tempfile.mkdtemp(prefix="duckdb_gen_")
        temp_source_path = Path(temp_dir) / f"source{file_extension}"
        temp_duckdb_path = Path(temp_dir) / "output.duckdb"

        try:
            # Write source data to temp file
            temp_source_path.write_bytes(binary_data)

            # Create DuckDB connection with configured settings
            conn = duckdb.connect(str(temp_duckdb_path))
            try:
                conn.execute(f"SET memory_limit = '{settings.duckdb_memory_limit}'")
                conn.execute(f"SET temp_directory = '{settings.duckdb_temp_dir}'")
                conn.execute("SET preserve_insertion_order = false")

                # Import data based on file type
                if file_extension == ".xlsx":
                    self._import_xlsx(conn, temp_source_path, source_file)
                elif file_extension == ".xls":
                    self._import_xls(conn, temp_source_path, source_file)
                elif file_extension in (".csv", ".tsv"):
                    self._import_csv(conn, temp_source_path, source_file)
                else:
                    raise ValueError(f"Unsupported extension: {file_extension}")

                # Step 5: Run SUMMARIZE on all tables
                summary = self._summarize_tables(conn)

                # Step 6: Extract sample data
                sample_data = self._extract_sample_data(conn)

                # Step 7: Build table info
                tables = self._build_table_info(conn)

                # Step 8: CHECKPOINT to merge WAL
                conn.execute("CHECKPOINT")

            finally:
                conn.close()

            # Read the generated DuckDB file
            duckdb_bytes = temp_duckdb_path.read_bytes()

            return DuckDBGenerateResult(
                duckdb_bytes=duckdb_bytes,
                summary=summary,
                tables=tables,
                sample_data=sample_data,
            )

        except Exception as exc:
            logger.error("DuckDB generation failed: %s", exc)
            raise RuntimeError(f"DuckDB generation failed: {exc}") from exc
        finally:
            # Cleanup temp files
            self._cleanup_temp_dir(temp_dir)

    def _import_xlsx(
        self, conn: duckdb.DuckDBPyConnection, source_path: Path, source_file: str
    ) -> None:
        """Import .xlsx file using DuckDB's excel extension.

        Args:
            conn: Active DuckDB connection.
            source_path: Path to the temporary source file.
            source_file: Original filename for table naming.
        """
        conn.execute("INSTALL excel")
        conn.execute("LOAD excel")

        # Read sheet names from the Excel file
        sheet_names = self._get_xlsx_sheet_names(source_path)

        if not sheet_names:
            raise RuntimeError("No sheets found in the Excel file")

        base_name = self._extract_table_name(source_file)
        safe_path = str(source_path).replace("'", "\\'")

        if len(sheet_names) == 1:
            # Single sheet: use filename as table name
            table_name = self._sanitize_table_name(base_name)
            conn.execute(
                f'CREATE TABLE "{table_name}" AS '
                f"SELECT * FROM read_xlsx('{safe_path}', header=true)"
            )
            logger.info("Imported single sheet as table '%s'", table_name)
        else:
            # Multi-sheet: import each non-empty sheet
            used_names: list[str] = []
            for sheet_name in sheet_names:
                safe_sheet = sheet_name.replace("'", "\\'")
                table_name = self._sanitize_table_name(f"sheet_{sheet_name}")
                table_name = self._deduplicate_table_name(table_name, used_names)
                used_names.append(table_name)

                try:
                    conn.execute(
                        f'CREATE TABLE "{table_name}" AS '
                        f"SELECT * FROM read_xlsx('{safe_path}', "
                        f"header=true, sheet='{safe_sheet}')"
                    )
                    # Check if the table has any rows; if empty, drop it
                    count = conn.execute(
                        f'SELECT COUNT(*) FROM "{table_name}"'
                    ).fetchone()
                    if count is not None and count[0] == 0:
                        conn.execute(f'DROP TABLE "{table_name}"')
                        logger.info(
                            "Skipped empty sheet '%s' (table '%s')",
                            sheet_name,
                            table_name,
                        )
                        used_names.remove(table_name)
                    else:
                        logger.info(
                            "Imported sheet '%s' as table '%s'",
                            sheet_name,
                            table_name,
                        )
                except Exception as exc:
                    logger.warning("Failed to import sheet '%s': %s", sheet_name, exc)

    def _import_xls(
        self, conn: duckdb.DuckDBPyConnection, source_path: Path, source_file: str
    ) -> None:
        """Import legacy .xls file using openpyxl as best-effort fallback.

        Note: openpyxl only supports .xlsx format (Office Open XML).
        Legacy .xls files (Binary Excel format) require xlrd. If openpyxl
        fails to read the file, a clear error is raised.

        Args:
            conn: Active DuckDB connection.
            source_path: Path to the temporary source file.
            source_file: Original filename for table naming.
        """
        try:
            import openpyxl
        except ImportError as exc:
            raise RuntimeError(
                "openpyxl is required for .xls file support. "
                "Install it with: pip install openpyxl"
            ) from exc

        try:
            wb = openpyxl.load_workbook(
                str(source_path), read_only=True, data_only=True
            )
        except Exception as exc:
            raise RuntimeError(
                f"Failed to read .xls file. openpyxl only supports .xlsx format. "
                f"For legacy .xls files, please convert to .xlsx first. Error: {exc}"
            ) from exc
        try:
            base_name = self._extract_table_name(source_file)
            used_names: list[str] = []

            sheet_names = wb.sheetnames
            if not sheet_names:
                raise RuntimeError("No sheets found in the .xls file")

            for sheet_name in sheet_names:
                ws = wb[sheet_name]

                # Read sheet data into rows
                rows = list(ws.iter_rows(values_only=True))
                if not rows or len(rows) < 2:
                    # Skip empty sheets or sheets with only headers
                    continue

                # First row as headers
                headers = [
                    str(h) if h is not None else f"column_{i}"
                    for i, h in enumerate(rows[0])
                ]
                data_rows = rows[1:]

                # Convert to list of dicts for DataFrame
                records = []
                for row in data_rows:
                    record = {}
                    for i, (header, value) in enumerate(zip(headers, row, strict=True)):
                        record[header] = value
                    records.append(record)

                if not records:
                    continue

                # Create DataFrame and register as DuckDB table
                import pandas as pd

                df = pd.DataFrame(records)

                if len(sheet_names) == 1:
                    table_name = self._sanitize_table_name(base_name)
                else:
                    table_name = self._sanitize_table_name(f"sheet_{sheet_name}")

                table_name = self._deduplicate_table_name(table_name, used_names)
                used_names.append(table_name)

                conn.execute(f'CREATE TABLE "{table_name}" AS SELECT * FROM df')
                logger.info(
                    "Imported .xls sheet '%s' as table '%s'", sheet_name, table_name
                )

        finally:
            wb.close()

    def _import_csv(
        self, conn: duckdb.DuckDBPyConnection, source_path: Path, source_file: str
    ) -> None:
        """Import CSV/TSV file using DuckDB's read_csv_auto.

        Args:
            conn: Active DuckDB connection.
            source_path: Path to the temporary source file.
            source_file: Original filename for table naming.
        """
        base_name = self._extract_table_name(source_file)
        table_name = self._sanitize_table_name(base_name)
        safe_path = str(source_path).replace("'", "\\'")

        conn.execute(
            f'CREATE TABLE "{table_name}" AS '
            f"SELECT * FROM read_csv_auto('{safe_path}')"
        )
        logger.info("Imported CSV/TSV as table '%s'", table_name)

    def _get_xlsx_sheet_names(self, source_path: Path) -> list[str]:
        """Get sheet names from an .xlsx file using openpyxl.

        Args:
            source_path: Path to the Excel file.

        Returns:
            List of sheet names.
        """
        try:
            import openpyxl

            wb = openpyxl.load_workbook(str(source_path), read_only=True)
            try:
                return wb.sheetnames
            finally:
                wb.close()
        except Exception as exc:
            logger.warning("Failed to read sheet names with openpyxl: %s", exc)
            # Fallback: return empty, DuckDB will import the default sheet
            return []

    def _summarize_tables(self, conn: duckdb.DuckDBPyConnection) -> dict[str, Any]:
        """Run SUMMARIZE on all tables and collect results.

        Args:
            conn: Active DuckDB connection.

        Returns:
            Dict mapping table name to SUMMARIZE results.
        """
        summary: dict[str, Any] = {}
        tables = conn.execute(
            "SELECT table_name FROM information_schema.tables "
            "WHERE table_schema = 'main'"
        ).fetchall()

        for (table_name,) in tables:
            try:
                result = conn.execute(f'SUMMARIZE "{table_name}"').fetchall()
                columns = [
                    "column_name",
                    "column_type",
                    "min",
                    "max",
                    "approx_unique",
                    "avg",
                    "std",
                    "q25",
                    "q50",
                    "q75",
                    "count",
                    "null_percentage",
                ]
                table_summary = []
                for row in result:
                    row_dict = {}
                    for i, col in enumerate(columns):
                        if i < len(row):
                            value = row[i]
                            # Convert non-serializable types to string
                            if value is not None and not isinstance(
                                value, (str, int, float, bool)
                            ):
                                value = str(value)
                            row_dict[col] = value
                    table_summary.append(row_dict)
                summary[table_name] = table_summary
                logger.debug("SUMMARIZE completed for table '%s'", table_name)
            except Exception as exc:
                logger.warning("SUMMARIZE failed for table '%s': %s", table_name, exc)
                summary[table_name] = {"error": str(exc)}

        return summary

    def _extract_sample_data(
        self, conn: duckdb.DuckDBPyConnection
    ) -> dict[str, list[Any]]:
        """Extract sample rows from all tables.

        Args:
            conn: Active DuckDB connection.

        Returns:
            Dict mapping table name to list of row dicts.
        """
        sample_data: dict[str, list[Any]] = {}
        sample_rows = self._settings.duckdb_summary_sample_rows

        tables = conn.execute(
            "SELECT table_name FROM information_schema.tables "
            "WHERE table_schema = 'main'"
        ).fetchall()

        for (table_name,) in tables:
            try:
                result = conn.execute(
                    f'SELECT * FROM "{table_name}" LIMIT {sample_rows}'
                )
                columns = [desc[0] for desc in result.description]
                rows = result.fetchall()

                table_samples = []
                for row_idx, row in enumerate(rows):
                    row_dict = {}
                    try:
                        for col, val in zip(columns, row, strict=True):
                            # Convert non-serializable types to string
                            if val is not None and not isinstance(
                                val, (str, int, float, bool)
                            ):
                                val = str(val)
                            row_dict[col] = val
                    except ValueError as exc:
                        raise ValueError(
                            f"Column/row length mismatch in table '{table_name}' "
                            f"at row {row_idx}: expected {len(columns)} columns, "
                            f"got {len(row)} values"
                        ) from exc
                    table_samples.append(row_dict)

                sample_data[table_name] = table_samples
                logger.debug(
                    "Extracted %d sample rows from table '%s'",
                    len(table_samples),
                    table_name,
                )
            except Exception as exc:
                logger.warning(
                    "Failed to extract sample data from table '%s': %s",
                    table_name,
                    exc,
                )
                sample_data[table_name] = []

        return sample_data

    def _build_table_info(
        self, conn: duckdb.DuckDBPyConnection
    ) -> list[DuckDBTableInfo]:
        """Build table info with column metadata for all tables.

        Args:
            conn: Active DuckDB connection.

        Returns:
            List of DuckDBTableInfo for each table.
        """
        tables_info: list[DuckDBTableInfo] = []

        tables = conn.execute(
            "SELECT table_name FROM information_schema.tables "
            "WHERE table_schema = 'main'"
        ).fetchall()

        for (table_name,) in tables:
            try:
                # Get row count
                count_result = conn.execute(
                    f'SELECT COUNT(*) FROM "{table_name}"'
                ).fetchone()
                row_count = count_result[0] if count_result else 0

                # Get column info using parameterized query
                columns_info: list[DuckDBColumnInfo] = []
                col_result = conn.execute(
                    "SELECT column_name, data_type, is_nullable "
                    "FROM information_schema.columns "
                    "WHERE table_name = ? AND table_schema = 'main' "
                    "ORDER BY ordinal_position",
                    [table_name],
                ).fetchall()

                for col_name, col_type, _is_nullable in col_result:
                    # Get null count for this column
                    try:
                        null_count_result = conn.execute(
                            f'SELECT COUNT(*) FROM "{table_name}" '
                            f'WHERE "{col_name}" IS NULL'
                        ).fetchone()
                        null_count = null_count_result[0] if null_count_result else 0
                    except Exception:
                        null_count = 0

                    columns_info.append(
                        DuckDBColumnInfo(
                            name=col_name,
                            type=col_type,
                            null_count=null_count,
                        )
                    )

                tables_info.append(
                    DuckDBTableInfo(
                        name=table_name,
                        row_count=row_count,
                        columns=columns_info,
                    )
                )
            except Exception as exc:
                logger.warning(
                    "Failed to build info for table '%s': %s", table_name, exc
                )

        return tables_info

    def _extract_table_name(self, source_file: str) -> str:
        """Extract a base table name from the source filename.

        Args:
            source_file: Original filename.

        Returns:
            Base name without extension.
        """
        name = Path(source_file).stem
        if not name:
            name = "data"
        return name

    def _sanitize_table_name(self, name: str) -> str:
        """Replace special characters in a table name with underscores.

        Args:
            name: Raw table name.

        Returns:
            Sanitized table name safe for DuckDB identifiers.
        """
        # Replace non-alphanumeric characters (except underscores) with underscores
        sanitized = re.sub(r"[^a-zA-Z0-9_]", "_", name)
        # Collapse multiple underscores
        sanitized = re.sub(r"_+", "_", sanitized)
        # Strip leading/trailing underscores
        sanitized = sanitized.strip("_")
        # Ensure the name starts with a letter or underscore
        if sanitized and sanitized[0].isdigit():
            sanitized = f"t_{sanitized}"
        # Fallback if empty
        if not sanitized:
            sanitized = "data"
        return sanitized.lower()

    def _deduplicate_table_name(self, name: str, existing_names: list[str]) -> str:
        """Ensure table name is unique by appending _1, _2, etc. if needed.

        Args:
            name: Desired table name.
            existing_names: List of already-used table names.

        Returns:
            Unique table name.
        """
        if name not in existing_names:
            return name

        counter = 1
        while f"{name}_{counter}" in existing_names:
            counter += 1
        return f"{name}_{counter}"

    def _cleanup_temp_dir(self, temp_dir: str) -> None:
        """Clean up temporary directory and its contents.

        Args:
            temp_dir: Path to the temporary directory.
        """
        try:
            shutil.rmtree(temp_dir, ignore_errors=True)
        except Exception as exc:
            logger.warning("Failed to cleanup temp dir '%s': %s", temp_dir, exc)
